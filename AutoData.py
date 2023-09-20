import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
import pandas as pd
import threading
from openpyxl import load_workbook
from tkinter import Listbox

# DataFrame que irá armazenar todas as colunas e linhas
dataframe = None
root = None
arquivo_excel = None
sheets = None

def escolher_arquivo():
    global dataframe
    arquivo = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
    if arquivo:
        workbook = load_workbook(arquivo_excel, read_only=True)
        sheets = workbook.sheetnames
        if len(sheets) == 1:
            selecionar_colunas_para_substituir()
        else:
            mostrar_lista_sheets()
        # Carregar o arquivo Excel selecionado em um DataFrame
        dataframe = pd.read_excel(arquivo)
        mostrar_botao_segundo_arquivo()

def mostrar_botao_segundo_arquivo():
    escolher_button.pack_forget()  # Remover o botão de escolher a base primária
    selecionar_button.pack()  # Mostrar o botão para selecionar o segundo arquivo

def selecionar_segundo_arquivo():
    global arquivo_excel, sheets
    arquivo_excel = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
    if arquivo_excel:
        # Carregar as planilhas (sheets) do arquivo Excel selecionado
        workbook = load_workbook(arquivo_excel, read_only=True)
        sheets = workbook.sheetnames
        if len(sheets) == 1:
            selecionar_colunas_para_substituir()
        else:
            mostrar_lista_sheets()

def mostrar_lista_sheets():
    global root
    root.withdraw()  # Esconder a janela principal
    sheet_selection_window = tk.Tk()
    sheet_selection_window.title("Selecione as Sheets")
    sheet_selection_window.geometry("300x200")

    sheet_listbox = Listbox(sheet_selection_window, selectmode=tk.MULTIPLE)
    for sheet in sheets:
        sheet_listbox.insert(tk.END, sheet)
    sheet_listbox.pack()

    def confirmar_selecao():
        selected_sheets = [sheets[i] for i in sheet_listbox.curselection()]
        sheet_selection_window.destroy()
        selecionar_colunas_para_substituir(selected_sheets)

    confirm_button = tk.Button(sheet_selection_window, text="Confirmar", command=confirmar_selecao)
    confirm_button.pack()

def selecionar_colunas_para_substituir(selected_sheets=None):
    if selected_sheets is None:
        selected_sheets = sheets

    global root, dataframe, arquivo_excel

    root.withdraw()  # Esconder a janela principal
    coluna_selection_window = tk.Tk()
    coluna_selection_window.title("Selecione as Colunas para Substituir")
    coluna_selection_window.geometry("400x400")

    coluna_listbox = Listbox(coluna_selection_window, selectmode=tk.MULTIPLE)
    for coluna in dataframe.columns:
        coluna_listbox.insert(tk.END, coluna)
    coluna_listbox.pack()

    def confirmar_substituicao():
        selected_columns = [coluna_listbox.get(i) for i in coluna_listbox.curselection()]

        # Processar cada sheet selecionada
        for sheet_name in selected_sheets:
            # Ler o arquivo Excel selecionado em um DataFrame
            dataframe_excel = pd.read_excel(arquivo_excel, sheet_name=sheet_name)

            # Substituir colunas correspondentes no DataFrame Excel
            for coluna in selected_columns:
                if coluna in dataframe_excel.columns:
                    dataframe_excel[coluna] = dataframe[coluna]

            # Salvar o DataFrame Excel resultante em um novo arquivo
            arquivo_salvar = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")])
            if arquivo_salvar:
                with pd.ExcelWriter(arquivo_salvar, engine='openpyxl') as writer:
                    writer.book = load_workbook(arquivo_excel)
                    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
                    dataframe_excel.to_excel(writer, index=False, sheet_name=sheet_name)

        coluna_selection_window.destroy()
        root.deiconify()  # Mostrar a janela principal
        messagebox.showinfo("Concluído", "Análise e Modificações concluídas com sucesso!👍")

    confirm_button = tk.Button(coluna_selection_window, text="Confirmar Substituição", command=confirmar_substituicao)
    confirm_button.pack()

if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1280x720")  # Definir o tamanho da janela principal

    escolher_button = tk.Button(root, text="Escolha a Base Primaria", command=escolher_arquivo)
    escolher_button.pack()

    selecionar_button = tk.Button(root, text="Selecionar Segundo Arquivo", command=selecionar_segundo_arquivo)

    root.mainloop()
